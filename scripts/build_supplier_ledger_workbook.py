from __future__ import annotations

import re
import urllib.request
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "operations"
EXPORT_DATE = "2026-05-26"

SOURCE_ID = "1O3lM52KoombirF657zMMEJhFdYEqXOCKXsIrtgIWLwA"
SOURCE_URL = f"https://docs.google.com/spreadsheets/d/{SOURCE_ID}/export?format=xlsx"
SOURCE_XLSX = OUT_DIR / f"minhhong-sheet-goc-export-{EXPORT_DATE}.xlsx"

MANUAL_SUPPLIER_ID = "1JHIFHgbUnTcDCqqysmh6D6DN8fMqQl5tUNdI7uVsoOw"
MANUAL_SUPPLIER_URL = f"https://docs.google.com/spreadsheets/d/{MANUAL_SUPPLIER_ID}/export?format=xlsx"
MANUAL_SUPPLIER_XLSX = OUT_DIR / f"minhhong-sheet-moi-export-{EXPORT_DATE}.xlsx"

OUTPUT_XLSX = OUT_DIR / f"minhhong-cong-no-doi-tac-copy-{EXPORT_DATE}.xlsx"

GENERATED_SHEETS = [
    "Hướng dẫn",
    "Tổng quan",
    "Công nợ đối tác",
    "Đối chiếu Long",
    "Nhập hàng",
    "Thanh toán",
    "Trả hàng",
    "Đơn khách",
    "Đối tác",
    "Cảnh báo dữ liệu",
]

MONEY_RE = re.compile(r"[-+]?\d[\d\s.,]*")
PHONE_RE = re.compile(r"(?:0|84|\+84)[\d\s.\-]{7,}")


def clean(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def parse_money(value: object) -> int:
    if value is None or value == "":
        return 0
    if isinstance(value, (int, float)):
        return int(round(value))

    matches = list(MONEY_RE.finditer(str(value).strip()))
    if not matches:
        return 0

    amounts: list[int] = []
    for match in matches:
        raw = match.group(0)
        digits = re.sub(r"\D", "", raw)
        if digits:
            sign = -1 if raw.strip().startswith("-") else 1
            amounts.append(sign * int(digits))

    if not amounts:
        return 0
    return max(amounts, key=abs)


def parse_quantity(value: object) -> object:
    if value is None or value == "":
        return ""
    if isinstance(value, (int, float)):
        return int(value) if float(value).is_integer() else value

    text = str(value).strip()
    number = re.sub(r"[^\d.,]", "", text)
    if not number:
        return text

    try:
        parsed = float(number.replace(",", "."))
    except ValueError:
        return text

    return int(parsed) if parsed.is_integer() else parsed


def find_phone(*values: object) -> str:
    combined = " ".join(clean(value) for value in values if clean(value))
    match = PHONE_RE.search(combined)
    if match:
        digits = re.sub(r"\D", "", match.group(0))
        if digits.startswith("84") and len(digits) >= 10:
            digits = f"0{digits[2:]}"
        return digits

    digit_runs = re.findall(r"\d{9,10}", combined)
    for digits in digit_runs:
        if len(digits) == 10 and digits.startswith("0"):
            return digits
        if len(digits) == 9:
            return f"0{digits}"

    return ""


def infer_cell_quantity(item_name: str, amount: int) -> tuple[object, str, object]:
    match = re.search(r"(\d+)\s*cell", item_name, re.IGNORECASE)
    if not match:
        return "", "", ""

    quantity = int(match.group(1))
    unit_price = amount / quantity if quantity else ""
    if isinstance(unit_price, float) and unit_price.is_integer():
        unit_price = int(unit_price)
    return quantity, "cell", unit_price


def append_row(ws, row_idx: int, values: list[Any]) -> None:
    for col_idx, value in enumerate(values, 1):
        ws.cell(row_idx, col_idx, value)


def set_headers(ws, headers: list[str]) -> None:
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F2937")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def style_sheet(ws, max_col: int | None = None) -> None:
    max_col = max_col or ws.max_column
    thin = Side(style="thin", color="E5E7EB")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=max_col):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for col in range(1, max_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16

    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 28


def set_money_format(ws, cols: list[int], start: int = 2, end: int = 500) -> None:
    for col in cols:
        for row in range(start, end + 1):
            ws.cell(row=row, column=col).number_format = '#,##0 "đ"'


def add_dropdown(ws, formula: str, cell_range: str, allow_blank: bool = True) -> None:
    validation = DataValidation(type="list", formula1=formula, allow_blank=allow_blank)
    ws.add_data_validation(validation)
    validation.add(cell_range)


def write_note(ws, row: int, title: str, body: str) -> None:
    ws.cell(row=row, column=1, value=title).font = Font(bold=True, color="111827")
    ws.cell(row=row, column=2, value=body).alignment = Alignment(wrap_text=True, vertical="top")


def is_raw_purchase_summary_or_note_row(item: str, unit_or_label: str) -> bool:
    text = f"{item} {unit_or_label}".casefold()
    markers = ["tổng", "tong", "ck cho long", "đã tính", "da tinh", "còn nợ", "con no"]
    return any(marker in text for marker in markers)


def extract_raw_supplier_totals(wb_values) -> dict[str, int]:
    purchase_total = 0
    purchase_rows = 0
    excluded_total = 0

    if "Sheet1" in wb_values.sheetnames:
        source = wb_values["Sheet1"]
        for row in range(4, source.max_row + 1):
            item = clean(source.cell(row, 2).value)
            unit_or_label = clean(source.cell(row, 4).value)
            amount = parse_money(source.cell(row, 7).value)
            if not any([item, unit_or_label, amount]):
                continue
            if is_raw_purchase_summary_or_note_row(item, unit_or_label):
                excluded_total += amount
                continue
            purchase_total += amount
            purchase_rows += 1

    payment_total = 0
    if "Tiền chuyển cho long" in wb_values.sheetnames:
        payment_total = parse_money(wb_values["Tiền chuyển cho long"].cell(1, 3).value)

    return_total = 0
    if "Đơn trả lại" in wb_values.sheetnames:
        source = wb_values["Đơn trả lại"]
        for row in range(4, source.max_row + 1):
            return_total += parse_money(source.cell(row, 5).value)

    return {
        "raw_purchase_total": purchase_total,
        "raw_purchase_rows": purchase_rows,
        "raw_excluded_total": excluded_total,
        "raw_payment_total": payment_total,
        "raw_return_total": return_total,
        "raw_remaining": purchase_total - payment_total - return_total,
    }


def extract_manual_supplier_summary(wb_values) -> dict[str, Any]:
    detail = wb_values["Đơn hàng nhập từ long"] if "Đơn hàng nhập từ long" in wb_values.sheetnames else None
    summary = wb_values["Đơn hàng mua từ long"] if "Đơn hàng mua từ long" in wb_values.sheetnames else None

    result: dict[str, Any] = {
        "opening_note": "",
        "opening_balance": 0,
        "extra_items": [],
        "summary_date": "",
        "summary_total": 0,
        "summary_paid": 0,
        "summary_remaining": 0,
    }

    if detail is not None:
        result["opening_note"] = clean(detail.cell(1, 1).value)
        result["opening_balance"] = parse_money(detail.cell(1, 1).value)

        for row in range(2, detail.max_row + 1):
            name = clean(detail.cell(row, 1).value)
            amount = parse_money(detail.cell(row, 2).value)
            if name or amount:
                result["extra_items"].append(
                    {
                        "name": name,
                        "amount": amount,
                        "sheet_amount": amount,
                        "correction_note": "",
                        "source": f"Đơn hàng nhập từ long!A{row}:B{row}",
                    }
                )

    if summary is not None and summary.max_row >= 2:
        result["summary_date"] = summary.cell(2, 1).value
        result["summary_total"] = parse_money(summary.cell(2, 4).value)
        result["summary_paid"] = parse_money(summary.cell(2, 5).value)
        result["summary_remaining"] = parse_money(summary.cell(2, 6).value)

    if result["summary_total"] and result["opening_balance"] and len(result["extra_items"]) == 1:
        expected_extra = result["summary_total"] - result["opening_balance"]
        item = result["extra_items"][0]
        if expected_extra > 0 and item["amount"] != expected_extra:
            item["sheet_amount"] = item["amount"]
            item["amount"] = expected_extra
            item["correction_note"] = (
                f"Sheet đang ghi {item['sheet_amount']:,} nhưng tổng chốt cho thấy phát sinh thực tế là "
                f"{expected_extra:,}."
            ).replace(",", ".")

    detail_total = result["opening_balance"] + sum(item["amount"] for item in result["extra_items"])
    result["detail_total"] = detail_total
    result["detail_remaining"] = detail_total - result["summary_paid"]
    result["summary_mismatch"] = detail_total - result["summary_total"] if result["summary_total"] else 0
    result["remaining_mismatch"] = (
        result["detail_remaining"] - result["summary_remaining"] if result["summary_remaining"] else 0
    )
    result["adjustment"] = result["summary_total"] - detail_total if result["summary_total"] else 0
    return result


def build_partner_rows(wb_values) -> list[list[str]]:
    rows = [["Long", "Đối tác công nợ chính hiện tại", "", "Đang theo dõi"]]
    source_sellers: list[str] = []

    if "Sheet1" in wb_values.sheetnames:
        source = wb_values["Sheet1"]
        for row in range(4, source.max_row + 1):
            seller = clean(source.cell(row, 9).value)
            note = clean(source.cell(row, 8).value)
            if seller and seller not in source_sellers:
                source_sellers.append(seller)
            if "SHOPEE" in note.upper() and "Shopee" not in source_sellers:
                source_sellers.append("Shopee")

    for seller in source_sellers:
        if seller != "Long":
            rows.append(
                [
                    seller,
                    "Nguồn mua hộ qua Long trong dữ liệu cũ; chỉ tính công nợ riêng nếu Minh Hồng tự mua trực tiếp",
                    "",
                    "Nguồn tham khảo/mở rộng sau",
                ]
            )

    rows.append(["Khác", "Dùng tạm khi chưa biết tên đối tác", "", "Tùy dùng"])
    return rows


def build_workbook() -> Path:
    OUT_DIR.mkdir(exist_ok=True)
    urllib.request.urlretrieve(SOURCE_URL, SOURCE_XLSX)
    urllib.request.urlretrieve(MANUAL_SUPPLIER_URL, MANUAL_SUPPLIER_XLSX)

    wb = load_workbook(SOURCE_XLSX, data_only=False)
    wb_values = load_workbook(SOURCE_XLSX, data_only=True)
    manual_values = load_workbook(MANUAL_SUPPLIER_XLSX, data_only=True)
    manual_summary = extract_manual_supplier_summary(manual_values)
    manual_summary.update(extract_raw_supplier_totals(wb_values))
    raw_remaining = manual_summary.get("raw_remaining", 0)
    extra_total = sum(item["amount"] for item in manual_summary.get("extra_items", []))
    manual_summary["raw_same_period_remaining"] = raw_remaining + extra_total - manual_summary.get("summary_paid", 0)
    manual_summary["opening_vs_raw_mismatch"] = manual_summary.get("opening_balance", 0) - raw_remaining
    manual_summary["same_period_mismatch"] = (
        manual_summary.get("summary_remaining", 0) - manual_summary["raw_same_period_remaining"]
    )

    for title in GENERATED_SHEETS:
        if title in wb.sheetnames:
            del wb[title]

    for index, title in enumerate(GENERATED_SHEETS):
        wb.create_sheet(title, index=index)

    ws_help = wb["Hướng dẫn"]
    ws_overview = wb["Tổng quan"]
    ws_debt = wb["Công nợ đối tác"]
    ws_reconcile = wb["Đối chiếu Long"]
    ws_purchase = wb["Nhập hàng"]
    ws_payment = wb["Thanh toán"]
    ws_return = wb["Trả hàng"]
    ws_sales = wb["Đơn khách"]
    ws_partner = wb["Đối tác"]
    ws_warning = wb["Cảnh báo dữ liệu"]

    partner_rows = build_partner_rows(wb_values)

    build_partner_sheet(ws_partner, partner_rows)
    build_purchase_sheet(ws_purchase, wb_values, manual_summary)
    build_payment_sheet(ws_payment, wb_values, manual_summary)
    build_return_sheet(ws_return, wb_values, manual_summary)
    build_sales_sheet(ws_sales, wb_values)
    summary_row = build_debt_sheet(ws_debt, partner_rows, manual_summary)
    build_overview_sheet(ws_overview, summary_row, manual_summary)
    build_long_reconciliation_sheet(ws_reconcile, manual_summary)
    build_help_sheet(ws_help)
    build_warning_sheet(ws_warning, manual_summary)
    add_validations(ws_purchase, ws_payment, ws_return)
    highlight_formula_cells([ws_debt, ws_overview, ws_sales])

    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except Exception:
        pass

    wb.save(OUTPUT_XLSX)
    return OUTPUT_XLSX


def build_partner_sheet(ws, partner_rows: list[list[str]]) -> None:
    headers = ["Tên đối tác", "Vai trò", "Số điện thoại", "Trạng thái/Ghi chú"]
    set_headers(ws, headers)

    for row_idx, row in enumerate(partner_rows, 2):
        append_row(ws, row_idx, row)

    style_sheet(ws, len(headers))
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 78
    ws.column_dimensions["D"].width = 28


def build_purchase_sheet(ws, wb_values, manual_summary: dict[str, Any]) -> None:
    headers = [
        "Mã nhập",
        "Ngày nhập",
        "Đối tác công nợ",
        "Người bán/nguồn gốc",
        "Tên hàng",
        "Loại",
        "Số lượng",
        "Đơn vị",
        "Đơn giá",
        "Thành tiền",
        "Đã nhận hàng",
        "Tính công nợ",
        "Ghi chú",
        "Dòng gốc",
    ]
    set_headers(ws, headers)
    output_row = 2
    has_manual_balance = bool(manual_summary.get("opening_balance") or manual_summary.get("summary_total"))

    if has_manual_balance:
        opening_balance = manual_summary["opening_balance"]
        append_row(
            ws,
            output_row,
            [
                f"NH-{output_row - 1:04d}",
                "07/05/2026",
                "Long",
                "Chốt công nợ",
                "Nợ tạm tính đến 07/05/2026",
                "Số dư chốt",
                1,
                "lần",
                opening_balance,
                opening_balance,
                "Có",
                "Có",
                "Số dư do Minh Hồng chốt; các dòng mua cũ bên dưới giữ để đối chiếu nhưng không cộng đôi.",
                "Đơn hàng nhập từ long!A1",
            ],
        )
        output_row += 1

        for item in manual_summary["extra_items"]:
            qty, unit, unit_price = infer_cell_quantity(item["name"], item["amount"])
            item_note = "Phát sinh sau số dư chốt 07/05/2026."
            if item.get("correction_note"):
                item_note = f"Phát sinh sau số dư chốt 07/05/2026. {item['correction_note']}"
            append_row(
                ws,
                output_row,
                [
                    f"NH-{output_row - 1:04d}",
                    manual_summary["summary_date"] or "08/05/2026",
                    "Long",
                    "Long",
                    item["name"],
                    "Pin/cell",
                    qty,
                    unit,
                    unit_price,
                    item["amount"],
                    "Có",
                    "Có",
                    item_note,
                    item["source"],
                ],
            )
            output_row += 1

        if manual_summary["adjustment"]:
            append_row(
                ws,
                output_row,
                [
                    f"NH-{output_row - 1:04d}",
                    manual_summary["summary_date"] or "08/05/2026",
                    "Long",
                    "Đối chiếu",
                    "Điều chỉnh tạm theo tổng Minh Hồng ghi",
                    "Điều chỉnh",
                    1,
                    "lần",
                    manual_summary["adjustment"],
                    manual_summary["adjustment"],
                    "Có",
                    "Có",
                    "Chi tiết nợ chốt + phát sinh lệch với tổng ghi; dòng này tạm đưa công nợ khớp tổng chốt, cần xác nhận.",
                    "Đơn hàng mua từ long!D2",
                ],
            )
            output_row += 1

    if "Sheet1" in wb_values.sheetnames:
        source = wb_values["Sheet1"]
        for row in range(4, source.max_row + 1):
            item = clean(source.cell(row, 2).value)
            if not item:
                continue

            note = clean(source.cell(row, 8).value)
            seller = clean(source.cell(row, 9).value)
            source_seller = seller or ("Shopee" if "SHOPEE" in note.upper() else "")
            include_debt = "Không" if has_manual_balance else "Có"
            source_note = note
            if has_manual_balance:
                source_note = f"{note} | Đã gộp trong số dư chốt 07/05/2026".strip(" |")

            values = [
                f"NH-{output_row - 1:04d}",
                "",
                "Long",
                source_seller,
                item,
                "Vật tư/đồ nghề",
                parse_quantity(source.cell(row, 3).value),
                "",
                parse_money(source.cell(row, 4).value) or "",
                parse_money(source.cell(row, 7).value) or "",
                "Có",
                include_debt,
                source_note,
                f"Sheet1!A{row}:J{row}",
            ]

            append_row(ws, output_row, values)
            output_row += 1

    for row in range(output_row, max(output_row + 60, 140)):
        ws.cell(row, 1, f"NH-{row - 1:04d}")
        ws.cell(row, 3, "Long")
        ws.cell(row, 10, f'=IF(OR(G{row}="",I{row}=""),"",G{row}*I{row})')
        ws.cell(row, 11, "Có")
        ws.cell(row, 12, "Có")

    style_sheet(ws, len(headers))
    widths = {"A": 12, "B": 14, "C": 22, "D": 28, "E": 38, "L": 16, "M": 64, "N": 22}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    set_money_format(ws, [9, 10], 2, ws.max_row)


def build_payment_sheet(ws, wb_values, manual_summary: dict[str, Any]) -> None:
    headers = ["Mã thanh toán", "Ngày", "Đối tác", "Số tiền", "Phương thức", "Ghi chú", "Tính công nợ"]
    set_headers(ws, headers)
    output_row = 2
    has_manual_balance = bool(manual_summary.get("opening_balance") or manual_summary.get("summary_total"))

    if "Tiền chuyển cho long" in wb_values.sheetnames:
        source = wb_values["Tiền chuyển cho long"]
        note = clean(source.cell(1, 1).value)
        amount = parse_money(source.cell(1, 3).value)
        date_match = re.search(r"(\d{1,2}/\d{1,2}/\d{2,4})", note)
        date_text = date_match.group(1) if date_match else ""

        if amount:
            values = [
                f"TT-{output_row - 1:04d}",
                date_text,
                "Long",
                amount,
                "Chuyển khoản",
                f"{note} | Đã xác nhận số đúng là 45.000.000; đã gộp trong số dư chốt 07/05/2026"
                if has_manual_balance
                else note,
                "Không" if has_manual_balance else "Có",
            ]
            append_row(ws, output_row, values)
            output_row += 1

    if has_manual_balance and manual_summary["summary_paid"]:
        values = [
            f"TT-{output_row - 1:04d}",
            manual_summary["summary_date"] or "08/05/2026",
            "Long",
            manual_summary["summary_paid"],
            "Trả trước",
            "Khoản trả trước theo sheet mới sau khi chốt nợ 07/05/2026.",
            "Có",
        ]
        append_row(ws, output_row, values)
        output_row += 1

    for row in range(output_row, max(output_row + 60, 80)):
        ws.cell(row, 1, f"TT-{row - 1:04d}")
        ws.cell(row, 3, "Long")
        ws.cell(row, 7, "Có")

    style_sheet(ws, len(headers))
    ws.column_dimensions["F"].width = 78
    set_money_format(ws, [4], 2, ws.max_row)


def build_return_sheet(ws, wb_values, manual_summary: dict[str, Any]) -> None:
    headers = [
        "Mã trả",
        "Ngày",
        "Đối tác",
        "Tên hàng",
        "Loại",
        "Số lượng",
        "Đơn giá",
        "Thành tiền",
        "Tính công nợ",
        "Lý do/Ghi chú",
        "Dòng gốc",
    ]
    set_headers(ws, headers)
    output_row = 2
    has_manual_balance = bool(manual_summary.get("opening_balance") or manual_summary.get("summary_total"))

    if "Đơn trả lại" in wb_values.sheetnames:
        source = wb_values["Đơn trả lại"]
        for row in range(4, source.max_row + 1):
            item = clean(source.cell(row, 2).value)
            if not item:
                continue

            include_debt = "Không" if has_manual_balance else "Có"
            note = "Nhập từ tab Đơn trả lại"
            if has_manual_balance:
                note = "Đã gộp trong số dư chốt 07/05/2026; giữ lại để đối chiếu."

            values = [
                f"TR-{output_row - 1:04d}",
                "",
                "Long",
                item,
                clean(source.cell(row, 3).value),
                parse_quantity(source.cell(row, 4).value),
                "",
                parse_money(source.cell(row, 5).value) or "",
                include_debt,
                note,
                f"Đơn trả lại!A{row}:E{row}",
            ]

            append_row(ws, output_row, values)
            output_row += 1

    for row in range(output_row, max(output_row + 60, 80)):
        ws.cell(row, 1, f"TR-{row - 1:04d}")
        ws.cell(row, 3, "Long")
        ws.cell(row, 8, f'=IF(OR(F{row}="",G{row}=""),"",F{row}*G{row})')
        ws.cell(row, 9, "Có")

    style_sheet(ws, len(headers))
    for col, width in {"D": 34, "I": 16, "J": 54, "K": 18}.items():
        ws.column_dimensions[col].width = width
    set_money_format(ws, [7, 8], 2, ws.max_row)


def build_sales_sheet(ws, wb_values) -> None:
    headers = [
        "Mã đơn",
        "Ngày mua",
        "Tên khách",
        "Số điện thoại",
        "Sản phẩm",
        "Tổng tiền",
        "Đã thu",
        "Còn nợ",
        "Ghi chú",
        "Trạng thái dữ liệu",
        "Dòng gốc",
    ]
    set_headers(ws, headers)
    output_row = 2

    if "Đơn hàng đã bán" in wb_values.sheetnames:
        source = wb_values["Đơn hàng đã bán"]
        for row in range(4, source.max_row + 1):
            name = clean(source.cell(row, 1).value)
            product = clean(source.cell(row, 2).value)
            total = parse_money(source.cell(row, 4).value)
            paid = parse_money(source.cell(row, 5).value)
            source_remaining = parse_money(source.cell(row, 6).value)
            note = clean(source.cell(row, 7).value)
            date_text = clean(source.cell(row, 8).value)
            phone = find_phone(source.cell(row, 9).value, source.cell(row, 7).value, source.cell(row, 3).value, name)

            if not any([name, product, total, paid, source_remaining, note, date_text, phone]):
                continue

            data_status = "Quên giá" if not total and (name or product) else ""
            remaining_value: object
            if total or paid:
                remaining_value = f"=MAX(F{output_row}-G{output_row},0)"
            else:
                remaining_value = source_remaining or 0

            values = [
                f"DH-{output_row - 1:04d}",
                date_text,
                name,
                phone,
                product,
                total or "",
                paid or "",
                remaining_value,
                note,
                data_status,
                f"Đơn hàng đã bán!A{row}:K{row}",
            ]

            append_row(ws, output_row, values)
            output_row += 1

    for row in range(output_row, max(output_row + 80, 140)):
        ws.cell(row, 1, f"DH-{row - 1:04d}")
        ws.cell(row, 8, f"=MAX(F{row}-G{row},0)")

    style_sheet(ws, 13)
    for col, width in {"C": 24, "D": 18, "E": 38, "I": 34, "J": 18, "K": 24, "L": 18, "M": 18}.items():
        ws.column_dimensions[col].width = width
    set_money_format(ws, [6, 7, 8, 13], 2, ws.max_row)
    add_sales_summary(ws)


def add_sales_summary(ws) -> None:
    rows = [
        ("Tổng tiền:", "=SUM(F:F)", "A9D18E", "000000"),
        ("Tổng đã thu:", "=SUM(G:G)", "B7DEE8", "000000"),
        ("Nợ còn:", "=SUM(H:H)", "FF3300", "FFFFFF"),
    ]

    for row_idx, (label, formula, fill, font_color) in enumerate(rows, 2):
        ws.cell(row_idx, 12, label)
        ws.cell(row_idx, 13, formula)
        ws.cell(row_idx, 12).font = Font(bold=True)
        ws.cell(row_idx, 13).font = Font(bold=True, color=font_color)
        ws.cell(row_idx, 13).fill = PatternFill("solid", fgColor=fill)
        ws.cell(row_idx, 13).number_format = '#,##0 "đ"'


def build_debt_sheet(ws, partner_rows: list[list[str]], manual_summary: dict[str, Any]) -> int:
    headers = [
        "Đối tác",
        "Giá trị hàng lấy",
        "Đã thanh toán",
        "Hàng trả lại",
        "Công nợ còn phải trả",
        "Theo thô cùng mốc",
        "Chênh chốt - thô cùng mốc",
        "Trạng thái",
        "Ghi chú",
    ]
    set_headers(ws, headers)

    for row_idx, partner in enumerate([row[0] for row in partner_rows], 2):
        ws.cell(row_idx, 1, partner)
        ws.cell(
            row_idx,
            2,
            f'=SUMIFS(\'Nhập hàng\'!$J:$J,\'Nhập hàng\'!$C:$C,A{row_idx},\'Nhập hàng\'!$L:$L,"Có")',
        )
        ws.cell(
            row_idx,
            3,
            f'=SUMIFS(\'Thanh toán\'!$D:$D,\'Thanh toán\'!$C:$C,A{row_idx},\'Thanh toán\'!$G:$G,"Có")',
        )
        ws.cell(
            row_idx,
            4,
            f'=SUMIFS(\'Trả hàng\'!$H:$H,\'Trả hàng\'!$C:$C,A{row_idx},\'Trả hàng\'!$I:$I,"Có")',
        )
        ws.cell(row_idx, 5, f"=B{row_idx}-C{row_idx}-D{row_idx}")
        if partner == "Long" and manual_summary.get("raw_same_period_remaining") is not None:
            ws.cell(row_idx, 6, manual_summary.get("raw_same_period_remaining", 0))
            ws.cell(row_idx, 7, f"=E{row_idx}-F{row_idx}")
        ws.cell(
            row_idx,
            8,
            f'=IF(E{row_idx}>0,"Còn phải trả đối tác",IF(E{row_idx}<0,"Đã chuyển dư/tạm ứng","Đã cân bằng"))',
        )
        if partner == "Long" and manual_summary.get("summary_total"):
            ws.cell(
                row_idx,
                9,
                "Dùng số dư chốt 07/05/2026 + phát sinh 08/05/2026; cột thô cùng mốc chỉ để đối chiếu, không cộng đôi.",
            )

    summary_row = len(partner_rows) + 3
    ws.cell(summary_row, 1, "Tổng cộng").font = Font(bold=True)
    for col in range(2, 8):
        letter = get_column_letter(col)
        ws.cell(summary_row, col, f"=SUM({letter}2:{letter}{summary_row - 1})")
        ws.cell(summary_row, col).font = Font(bold=True)

    style_sheet(ws, len(headers))
    for col, width in {"A": 24, "B": 22, "C": 18, "D": 18, "E": 22, "F": 20, "G": 20, "H": 28, "I": 78}.items():
        ws.column_dimensions[col].width = width
    set_money_format(ws, [2, 3, 4, 5, 6, 7], 2, summary_row)
    ws.conditional_formatting.add(
        f"E2:E{summary_row}",
        CellIsRule(operator="greaterThan", formula=["0"], fill=PatternFill("solid", fgColor="FEE2E2")),
    )
    ws.conditional_formatting.add(
        f"E2:E{summary_row}",
        CellIsRule(operator="lessThan", formula=["0"], fill=PatternFill("solid", fgColor="DBEAFE")),
    )
    return summary_row


def build_overview_sheet(ws, summary_row: int, manual_summary: dict[str, Any]) -> None:
    ws["A1"] = "Tổng quan công nợ và đơn khách"
    ws["A1"].font = Font(size=16, bold=True, color="111827")
    ws.merge_cells("A1:D1")

    rows = [
        ("Tổng giá trị hàng/nợ tính công nợ", f"='Công nợ đối tác'!B{summary_row}"),
        ("Tổng đã thanh toán đối tác", f"='Công nợ đối tác'!C{summary_row}"),
        ("Tổng hàng trả lại", f"='Công nợ đối tác'!D{summary_row}"),
        ("Công nợ đối tác hiện tại", f"='Công nợ đối tác'!E{summary_row}"),
        ("Tổng tiền đơn khách", "=SUM('Đơn khách'!F:F)"),
        ("Đã thu từ khách", "=SUM('Đơn khách'!G:G)"),
        ("Khách còn nợ", "=SUM('Đơn khách'!H:H)"),
        ("Long: tổng theo chi tiết mới", manual_summary.get("detail_total", 0)),
        ("Long: tổng Minh Hồng ghi", manual_summary.get("summary_total", 0)),
        ("Long: lệch chi tiết - tổng ghi", manual_summary.get("summary_mismatch", 0)),
        ("Long: còn nợ thô trước 08/05", manual_summary.get("raw_remaining", 0)),
        ("Long: lệch chốt 07/05 - thô cũ", manual_summary.get("opening_vs_raw_mismatch", 0)),
        ("Long: còn nợ thô cùng mốc", manual_summary.get("raw_same_period_remaining", 0)),
        ("Long: chênh chính thức - thô cùng mốc", manual_summary.get("same_period_mismatch", 0)),
    ]

    for row_idx, (label, value) in enumerate(rows, 3):
        ws.cell(row_idx, 1, label)
        ws.cell(row_idx, 2, value)
        ws.cell(row_idx, 1).font = Font(bold=True)
        ws.cell(row_idx, 2).number_format = '#,##0 "đ"'

    ws["A17"] = "Cách hiểu công nợ"
    ws["A17"].font = Font(bold=True)
    ws["B17"] = (
        "Công nợ đối tác = Giá trị hàng/nợ tính công nợ - Đã thanh toán - Hàng trả lại. "
        "Dữ liệu cũ trước 07/05/2026 đã được gộp vào số dư chốt để tránh cộng đôi; số theo dòng thô chỉ dùng để đối chiếu cùng mốc thời gian."
    )
    ws["B17"].alignment = Alignment(wrap_text=True)
    style_sheet(ws, 4)
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 64


def build_long_reconciliation_sheet(ws, manual_summary: dict[str, Any]) -> None:
    ws["A1"] = "Đối chiếu công nợ Long"
    ws["A1"].font = Font(size=16, bold=True, color="111827")
    ws.merge_cells("A1:D1")

    extra_total = sum(item["amount"] for item in manual_summary.get("extra_items", []))
    sheet_extra_total = sum(item.get("sheet_amount", item["amount"]) for item in manual_summary.get("extra_items", []))
    extra_names = "; ".join(item["name"] for item in manual_summary.get("extra_items", []))
    rows = [
        ("Nợ tạm tính đến 07/05/2026", manual_summary.get("opening_balance", 0), manual_summary.get("opening_note", "")),
        (
            "Phát sinh sau chốt theo sheet",
            sheet_extra_total,
            extra_names,
        ),
        (
            "Phát sinh sau chốt sau sửa",
            extra_total,
            "Sửa theo tổng chốt; 300cell eve 25p tính là 7.490.000đ.",
        ),
        ("Tổng theo chi tiết", manual_summary.get("detail_total", 0), "Nợ chốt + phát sinh"),
        ("Tổng Minh Hồng ghi", manual_summary.get("summary_total", 0), "Theo tab Đơn hàng mua từ long"),
        ("Lệch chi tiết - tổng ghi", manual_summary.get("summary_mismatch", 0), "Cần xác nhận nếu khác 0"),
        ("Đã trả trước", manual_summary.get("summary_paid", 0), "Theo tab Đơn hàng mua từ long"),
        ("Còn nợ theo chi tiết", manual_summary.get("detail_remaining", 0), "Tổng theo chi tiết - đã trả trước"),
        ("Còn nợ theo tổng Minh Hồng ghi", manual_summary.get("summary_remaining", 0), "Tổng ghi - đã trả trước"),
        ("Hàng lấy theo dòng thô cũ", manual_summary.get("raw_purchase_total", 0), "Đã loại các dòng tổng/ghi chú để tránh cộng đôi trong phần thô."),
        ("Đã thanh toán theo dòng thô cũ", manual_summary.get("raw_payment_total", 0), "Số 45.000.000đ đã xác nhận."),
        ("Trả hàng theo dòng thô cũ", manual_summary.get("raw_return_total", 0), ""),
        ("Còn nợ thô trước 08/05", manual_summary.get("raw_remaining", 0), "66.445.500 - 45.000.000 - 1.500.000."),
        (
            "Chênh chốt 07/05 - thô cũ",
            manual_summary.get("opening_vs_raw_mismatch", 0),
            "20.230.000 - 19.945.500. Đây là so sánh cùng mốc trước phát sinh/trả tiền 08/05.",
        ),
        (
            "Còn nợ thô cùng mốc hiện tại",
            manual_summary.get("raw_same_period_remaining", 0),
            "19.945.500 + 7.490.000 - 15.000.000.",
        ),
        (
            "Chênh chính thức - thô cùng mốc",
            manual_summary.get("same_period_mismatch", 0),
            "12.720.000 - 12.435.500. Số chính thức vẫn theo số dư chốt 07/05/2026.",
        ),
    ]

    header_row = 3
    for col, header in enumerate(["Mục", "Số tiền", "Ghi chú"], 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F2937")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = "A3:C3"

    for row_idx, row in enumerate(rows, 4):
        append_row(ws, row_idx, list(row))

    conclusion_row = len(rows) + 6
    ws.cell(conclusion_row, 1, "Kết luận")
    ws.cell(conclusion_row, 1).font = Font(bold=True)
    if manual_summary.get("summary_mismatch"):
        ws.cell(
            conclusion_row,
            2,
            "Tổng và số còn nợ Minh Hồng ghi tự khớp, nhưng chi tiết đang lệch 10.000đ. Workbook tạm ưu tiên số tổng 12.720.000đ và ghi rõ dòng điều chỉnh.",
        )
    else:
        ws.cell(
            conclusion_row,
            2,
            "Chi tiết sau khi sửa phát sinh 300cell eve 25p còn 7.490.000đ đã khớp tổng chốt và còn nợ 12.720.000đ. Số 19.945.500đ là còn nợ thô trước 08/05; nếu đưa về cùng mốc hiện tại thì còn 12.435.500đ, lệch số chính thức 284.500đ.",
        )
    ws.cell(conclusion_row, 2).alignment = Alignment(wrap_text=True)

    style_sheet(ws, 3)
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 72
    set_money_format(ws, [2], 2, conclusion_row)


def build_help_sheet(ws) -> None:
    ws["A1"] = "Hướng dẫn dùng bản copy"
    ws["A1"].font = Font(size=16, bold=True)
    ws.merge_cells("A1:C1")

    rows = [
        ("Không sửa sheet gốc", "File này được tạo từ bản export của Google Sheet gốc và không ghi đè dữ liệu gốc."),
        (
            "Số dư chốt Long",
            "Nợ tạm tính đến 07/05/2026 được nhập như một dòng số dư chốt. Các dòng mua/trả hàng/thanh toán cũ vẫn giữ để đối chiếu nhưng mặc định không tính lại vào công nợ hiện tại.",
        ),
        (
            "Khi lấy hàng",
            "Nhập một dòng trong tab Nhập hàng, chọn Đối tác công nợ, nhập số lượng/đơn giá hoặc nhập trực tiếp Thành tiền, và để Tính công nợ = Có nếu phát sinh đó làm tăng nợ.",
        ),
        (
            "Khi thanh toán",
            "Nhập một dòng trong tab Thanh toán. Công nợ ở tab Công nợ đối tác sẽ tự giảm khi Tính công nợ = Có.",
        ),
        (
            "Khi trả hàng",
            "Nhập một dòng trong tab Trả hàng. Công nợ đối tác cũng tự giảm theo giá trị hàng trả khi Tính công nợ = Có.",
        ),
        (
            "Thêm đối tác mới",
            "Thêm tên vào tab Đối tác, rồi chọn tên đó ở Nhập hàng/Thanh toán/Trả hàng. Các nguồn cũ như Shopee/Lào Cai hiện chỉ là nguồn Long mua hộ, chưa phải công nợ riêng.",
        ),
        (
            "Đơn khách",
            "Tab Đơn khách giữ cả đơn thiếu số tiền vì Minh Hồng cần quản lý đơn và số lượng pin đã bán. Các đơn thiếu tiền hiện tại được gắn Trạng thái dữ liệu = Quên giá; ô tổng tiền/tổng đã thu/nợ còn nằm ở cột L:M.",
        ),
    ]

    for row_idx, (title, body) in enumerate(rows, 3):
        write_note(ws, row_idx, title, body)

    style_sheet(ws, 3)
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 96


def build_warning_sheet(ws, manual_summary: dict[str, Any]) -> None:
    headers = ["Mức", "Vấn đề", "Cách xử lý đề xuất"]
    set_headers(ws, headers)

    rows = [
        [
            "Đã xác nhận",
            "Ghi chú 50tr trước đây là nhầm; số đúng tại thời điểm đó là 45.000.000.",
            "Workbook mới giữ 45.000.000 trong lịch sử thanh toán cũ, nhưng không cộng đôi vì đã gộp vào số dư chốt 07/05/2026.",
        ],
        [
            "Quy tắc mới",
            "Shopee, Điện tử Lào Cai và các tên nguồn cũ là nơi Long mua hộ, không phải công nợ riêng của Minh Hồng tại thời điểm cũ.",
            "Tab Đối tác vẫn giữ các tên này như nguồn tham khảo/mở rộng sau; công nợ cũ mặc định vẫn thuộc Long.",
        ],
        [
            "Dữ liệu chốt",
            "Số dư chốt Long đến 07/05/2026 được ưu tiên hơn việc cộng lại toàn bộ dòng thô cũ.",
            "Các dòng cũ trong Nhập hàng/Thanh toán/Trả hàng có Tính công nợ = Không để tránh cộng đôi.",
        ],
        [
            "Đơn khách",
            "Một số đơn bán cũ thiếu số tiền bán/đã trả vì bị quên giá, nhưng vẫn là đơn có thật.",
            "Workbook vẫn giữ các dòng này trong tab Đơn khách với Trạng thái dữ liệu = Quên giá để quản lý lịch sử đơn và số lượng pin bán ra.",
        ],
        [
            "Bổ sung",
            "Tab Đơn khách đã thêm tổng tiền, tổng đã thu, nợ còn ở cột L:M để nhìn nhanh như sheet cũ.",
            "Khi mở bằng Excel/Google Sheets, công thức tổng sẽ tự tính lại.",
        ],
    ]

    if manual_summary.get("summary_mismatch"):
        rows.insert(
            1,
            [
                "Cần xác nhận",
                (
                    f"Nợ chốt {manual_summary['opening_balance']:,} + phát sinh "
                    f"{sum(item['amount'] for item in manual_summary['extra_items']):,} = "
                    f"{manual_summary['detail_total']:,}, nhưng tab tổng ghi {manual_summary['summary_total']:,}."
                ).replace(",", "."),
                (
                    "Workbook tạm thêm dòng điều chỉnh "
                    f"{manual_summary['adjustment']:,} để công nợ khớp số Minh Hồng ghi là "
                    f"{manual_summary['summary_remaining']:,}. Cần xác nhận lại lệch 10.000đ."
                ).replace(",", "."),
            ],
        )
    else:
        rows.insert(
            1,
            [
                "Đã xử lý",
                "Phát sinh 300cell eve 25p được sửa từ 7.500.000đ xuống 7.490.000đ theo tổng chốt, nên số còn nợ Long khớp 12.720.000đ.",
                "Không cần dòng điều chỉnh tạm; nếu sheet nguồn còn ghi 7.500.000đ thì workbook copy dùng số đã sửa để tính công nợ.",
            ],
        )

    if manual_summary.get("raw_remaining") is not None:
        rows.insert(
            3,
            [
                "Đối chiếu",
                "Workbook có cột Theo thô cùng mốc và Chênh chốt - thô cùng mốc để so sánh đúng thời điểm.",
                "Số 19.945.500đ là còn nợ thô trước 08/05; số chính thức sau này vẫn theo số dư chốt 07/05/2026.",
            ],
        )

    for row_idx, row in enumerate(rows, 2):
        append_row(ws, row_idx, row)

    style_sheet(ws, len(headers))
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 82
    ws.column_dimensions["C"].width = 86


def add_validations(ws_purchase, ws_payment, ws_return) -> None:
    for ws, cell_range in [(ws_purchase, "C2:C500"), (ws_payment, "C2:C500"), (ws_return, "C2:C500")]:
        add_dropdown(ws, "'Đối tác'!$A$2:$A$200", cell_range, allow_blank=False)

    add_dropdown(ws_purchase, '"Có,Chưa"', "K2:K500", allow_blank=True)
    add_dropdown(ws_purchase, '"Có,Không"', "L2:L500", allow_blank=False)
    add_dropdown(ws_payment, '"Có,Không"', "G2:G500", allow_blank=False)
    add_dropdown(ws_return, '"Có,Không"', "I2:I500", allow_blank=False)


def highlight_formula_cells(sheets) -> None:
    formula_fill = PatternFill("solid", fgColor="EFF6FF")
    for ws in sheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    cell.fill = formula_fill


if __name__ == "__main__":
    print(build_workbook())
