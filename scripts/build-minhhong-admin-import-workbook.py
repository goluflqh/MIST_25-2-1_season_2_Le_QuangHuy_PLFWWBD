from __future__ import annotations

import re
import unicodedata
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parents[1]
OPERATIONS_DIR = ROOT / "operations"
SOURCE_WORKBOOK = OPERATIONS_DIR / "minhhong-cong-no-doi-tac-copy-2026-05-26.xlsx"
OUTPUT_WORKBOOK = OPERATIONS_DIR / "minhhong-admin-import-template-2026-05-26.xlsx"

PARTNER_HEADERS = ["Mã đối tác", "Tên đối tác", "Loại", "Số điện thoại", "Ghi chú", "Trạng thái"]
PURCHASE_HEADERS = [
    "Mã nhập",
    "Ngày nhập",
    "Mã đối tác",
    "Tên đối tác",
    "Người bán/nguồn gốc",
    "Tên hàng",
    "Loại",
    "Số lượng",
    "Đơn vị",
    "Đơn giá",
    "Chiết khấu (%)",
    "Tiền chiết khấu",
    "Thành tiền",
    "Đã nhận hàng",
    "Tính công nợ",
    "Ghi chú",
    "Dòng gốc",
]
PAYMENT_HEADERS = [
    "Mã thanh toán",
    "Ngày",
    "Mã đối tác",
    "Tên đối tác",
    "Số tiền",
    "Phương thức",
    "Tính công nợ",
    "Ghi chú",
    "Dòng gốc",
]
RETURN_HEADERS = [
    "Mã trả",
    "Ngày",
    "Mã đối tác",
    "Tên đối tác",
    "Tên hàng",
    "Loại",
    "Số lượng",
    "Đơn giá",
    "Thành tiền",
    "Tính công nợ",
    "Lý do/Ghi chú",
    "Dòng gốc",
]
CUSTOMER_ORDER_HEADERS = [
    "Mã đơn",
    "Ngày mua",
    "Tên khách",
    "Số điện thoại",
    "Sản phẩm",
    "Tổng tiền",
    "Đã thu",
    "Còn nợ",
    "Trạng thái giá",
    "Ghi chú",
    "Dòng gốc",
]
RECONCILIATION_HEADERS = ["Khoá", "Nhãn", "Giá trị kỳ vọng", "Ghi chú"]

EXPECTED_LONG_PAYABLE = 12_720_000
EXPECTED_CUSTOMER_ORDER_ROWS = 41
EXPECTED_CUSTOMER_ORDER_TOTAL = 36_825_000
EXPECTED_CUSTOMER_ORDER_PAID = 29_790_000
EXPECTED_CUSTOMER_MISSING_PRICE_ROWS = 4


MONEY_FORMAT = '#,##0 "đ"'


def clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def money(value: Any) -> int:
    if value in (None, ""):
        return 0
    if isinstance(value, (int, float)):
        return int(round(value))
    digits = re.sub(r"[^0-9-]", "", str(value))
    if not digits or digits == "-":
        return 0
    return int(digits)


def number_or_blank(value: Any) -> Any:
    if value in (None, ""):
        return ""
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def yes(value: Any) -> bool:
    return clean(value).casefold() == "có"


def business_date(value: Any) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")

    text = clean(value)
    known_repairs = {
        "37/1/2026": "27/01/2026",
        "28/012026": "28/01/2026",
    }
    if text in known_repairs:
        return known_repairs[text]
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(text, fmt).strftime("%d/%m/%Y")
        except ValueError:
            pass
    return text


def slug_code(name: str, fallback_index: int) -> str:
    normalized = unicodedata.normalize("NFKD", name)
    ascii_name = normalized.encode("ascii", "ignore").decode("ascii")
    compact = re.sub(r"[^A-Za-z0-9]+", "_", ascii_name).strip("_").upper()
    if compact == "LONG":
        return "LONG"
    if compact == "KHAC":
        return "KHAC"
    if compact:
        return f"DT_{compact[:24]}"
    return f"DT_{fallback_index:03d}"


def set_headers(ws, headers: list[str]) -> None:
    for column, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=column, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F2937")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def append(ws, values: list[Any]) -> None:
    ws.append(values)


def style_sheet(ws, money_columns: set[int] | None = None) -> None:
    money_columns = money_columns or set()
    ws.sheet_view.showGridLines = False
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.row > 1 and cell.column in money_columns:
                cell.number_format = MONEY_FORMAT
    for column in range(1, ws.max_column + 1):
        width = 14
        for cell in ws.iter_cols(min_col=column, max_col=column, values_only=True):
            width = max(width, min(48, max((len(clean(value)) for value in cell if value is not None), default=0) + 2))
        ws.column_dimensions[get_column_letter(column)].width = width
    ws.row_dimensions[1].height = 28


def add_dropdown(ws, options: list[str], cell_range: str) -> None:
    validation = DataValidation(type="list", formula1='"' + ",".join(options) + '"', allow_blank=True)
    ws.add_data_validation(validation)
    validation.add(cell_range)


def read_partners(source_wb) -> tuple[list[list[Any]], dict[str, str]]:
    ws = source_wb["Đối tác"]
    partners: list[list[Any]] = []
    partner_codes: dict[str, str] = {}

    for row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 1):
        name = clean(row[0])
        if not name:
            continue
        role = clean(row[1])
        phone = clean(row[2])
        status = clean(row[3])
        code = slug_code(name, row_index)
        if code in partner_codes.values() and name not in partner_codes:
            code = f"{code}_{row_index:02d}"
        partner_codes[name] = code
        if name == "Long":
            partner_type = "Đối tác công nợ"
        elif name == "Khác":
            partner_type = "Dự phòng"
        else:
            partner_type = "Nguồn tham khảo"
        partners.append([code, name, partner_type, phone, role, status])

    if "Long" not in partner_codes:
        partner_codes["Long"] = "LONG"
        partners.insert(0, ["LONG", "Long", "Đối tác công nợ", "", "Đối tác công nợ chính hiện tại", "Đang theo dõi"])

    return partners, partner_codes


def partner_code(partner_codes: dict[str, str], partner_name: str) -> str:
    return partner_codes.get(partner_name) or slug_code(partner_name, len(partner_codes) + 1)


def build_purchase_rows(source_wb, partner_codes: dict[str, str]) -> tuple[list[list[Any]], dict[str, int]]:
    rows: list[list[Any]] = []
    totals = {"opening": 0, "counted_purchase": 0, "counted_return": 0, "reference_only": 0}
    ws = source_wb["Nhập hàng"]

    for source_row, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        amount = money(row[9])
        item_name = clean(row[4])
        if not amount or not item_name:
            continue

        partner_name = clean(row[2]) or "Long"
        category = clean(row[5])
        counts_flag = clean(row[11]) or "Có"
        if yes(counts_flag):
            if category == "Số dư chốt":
                totals["opening"] += amount
            else:
                totals["counted_purchase"] += amount
        else:
            totals["reference_only"] += amount

        rows.append(
            [
                clean(row[0]),
                business_date(row[1]),
                partner_code(partner_codes, partner_name),
                partner_name,
                clean(row[3]),
                item_name,
                category,
                number_or_blank(row[6]),
                clean(row[7]),
                money(row[8]) or "",
                "",
                0,
                amount,
                clean(row[10]) or "Có",
                counts_flag,
                clean(row[12]),
                clean(row[13]) or f"Nhập hàng!A{source_row}:N{source_row}",
            ]
        )

    return rows, totals


def build_payment_rows(source_wb, partner_codes: dict[str, str]) -> tuple[list[list[Any]], dict[str, int]]:
    rows: list[list[Any]] = []
    totals = {"counted_payment": 0, "historical_paid": 0, "reference_only": 0}
    ws = source_wb["Thanh toán"]

    for source_row, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        amount = money(row[3])
        if not amount:
            continue

        partner_name = clean(row[2]) or "Long"
        counts_flag = clean(row[6]) or "Có"
        totals["historical_paid"] += amount
        if yes(counts_flag):
            totals["counted_payment"] += amount
        else:
            totals["reference_only"] += amount

        rows.append(
            [
                clean(row[0]),
                business_date(row[1]),
                partner_code(partner_codes, partner_name),
                partner_name,
                amount,
                clean(row[4]),
                counts_flag,
                clean(row[5]),
                f"Thanh toán!A{source_row}:G{source_row}",
            ]
        )

    return rows, totals


def build_return_rows(source_wb, partner_codes: dict[str, str]) -> tuple[list[list[Any]], dict[str, int]]:
    rows: list[list[Any]] = []
    totals = {"counted_return": 0, "reference_only": 0}
    ws = source_wb["Trả hàng"]

    for source_row, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        amount = money(row[7])
        item_name = clean(row[3])
        if not amount or not item_name:
            continue

        partner_name = clean(row[2]) or "Long"
        counts_flag = clean(row[8]) or "Có"
        if yes(counts_flag):
            totals["counted_return"] += amount
        else:
            totals["reference_only"] += amount

        rows.append(
            [
                clean(row[0]),
                business_date(row[1]),
                partner_code(partner_codes, partner_name),
                partner_name,
                item_name,
                clean(row[4]),
                number_or_blank(row[5]),
                money(row[6]) or "",
                amount,
                counts_flag,
                clean(row[9]),
                clean(row[10]) or f"Trả hàng!A{source_row}:K{source_row}",
            ]
        )

    return rows, totals


def build_customer_order_rows(source_wb) -> tuple[list[list[Any]], dict[str, int]]:
    rows: list[list[Any]] = []
    totals = {"rows": 0, "priced_rows": 0, "total": 0, "paid": 0, "missing_price_rows": 0}
    ws = source_wb["Đơn khách"]

    for source_row, row in enumerate(ws.iter_rows(min_row=2, max_col=11, values_only=True), 2):
        order_code = clean(row[0])
        customer_name = clean(row[2])
        customer_phone = clean(row[3])
        product = clean(row[4])
        total = money(row[5])
        paid = money(row[6])
        note = clean(row[8])
        raw_source_ref = clean(row[10])
        has_business_content = any([clean(row[1]), customer_name, customer_phone, product, total, paid, note, clean(row[9]), raw_source_ref])
        if not has_business_content:
            continue

        source_ref = raw_source_ref or f"Đơn khách!A{source_row}:K{source_row}"
        price_status = clean(row[9])
        if not total and (customer_name or product):
            price_status = "Quên giá"
        remaining = max(total - paid, 0) if total or paid else money(row[7])

        rows.append(
            [
                order_code,
                business_date(row[1]),
                customer_name,
                customer_phone,
                product,
                total or "",
                paid or "",
                remaining,
                price_status,
                note,
                source_ref,
            ]
        )
        totals["rows"] += 1
        totals["total"] += total
        totals["paid"] += paid
        if total:
            totals["priced_rows"] += 1
        if price_status == "Quên giá":
            totals["missing_price_rows"] += 1

    return rows, totals


def assert_expected(totals: dict[str, int]) -> None:
    checks = {
        "Long payable": (totals["long_payable"], EXPECTED_LONG_PAYABLE),
        "Đơn khách meaningful rows": (totals["customer_order_rows"], EXPECTED_CUSTOMER_ORDER_ROWS),
        "Đơn khách total": (totals["customer_order_total"], EXPECTED_CUSTOMER_ORDER_TOTAL),
        "Đơn khách paid": (totals["customer_order_paid"], EXPECTED_CUSTOMER_ORDER_PAID),
        "Đơn khách missing price rows": (totals["customer_legacy_missing_price_rows"], EXPECTED_CUSTOMER_MISSING_PRICE_ROWS),
    }
    failures = [f"{name}: got {actual}, expected {expected}" for name, (actual, expected) in checks.items() if actual != expected]
    if failures:
        raise SystemExit("Workbook source does not match approved totals:\n" + "\n".join(failures))


def build_reconciliation_rows(totals: dict[str, int]) -> list[list[Any]]:
    return [
        ["long_opening_balance", "Long - số dư chốt 07/05/2026", totals["long_opening_balance"], "Tính vào công nợ hiện tại"],
        ["long_counted_purchase", "Long - phát sinh mua sau chốt", totals["long_counted_purchase"], "Không gồm số dư chốt"],
        ["long_counted_payment", "Long - đã thanh toán sau chốt", totals["long_counted_payment"], "Giảm công nợ hiện tại"],
        ["long_payable", "Long - Minh Hồng cần trả", totals["long_payable"], "Phải bằng 12.720.000 trước VPS"],
        ["long_historical_paid", "Long - đã thanh toán lịch sử", totals["long_historical_paid"], "Gồm 45.000.000 tham khảo + 15.000.000 hiện tại"],
        ["customer_order_rows", "Đơn khách - số dòng nghiệp vụ", totals["customer_order_rows"], "Không gồm các dòng DH-* rỗng"],
        ["customer_order_total", "Đơn khách - tổng tiền", totals["customer_order_total"], "37 dòng có giá"],
        ["customer_order_paid", "Đơn khách - đã thu", totals["customer_order_paid"], "Tổng tiền đã thu từ khách"],
        ["customer_legacy_missing_price_rows", "Đơn khách - dòng cũ quên giá", totals["customer_legacy_missing_price_rows"], "Giữ để theo dõi, không tự đoán giá"],
    ]


def write_sheet(wb: Workbook, title: str, headers: list[str], rows: list[list[Any]], money_columns: set[int] | None = None) -> None:
    ws = wb.create_sheet(title)
    set_headers(ws, headers)
    for row in rows:
        append(ws, row)
    style_sheet(ws, money_columns)


def build_workbook() -> Path:
    if not SOURCE_WORKBOOK.exists():
        raise SystemExit(f"Missing source workbook: {SOURCE_WORKBOOK}")

    source_wb = load_workbook(SOURCE_WORKBOOK, data_only=True)
    partners, partner_codes = read_partners(source_wb)
    purchases, purchase_totals = build_purchase_rows(source_wb, partner_codes)
    payments, payment_totals = build_payment_rows(source_wb, partner_codes)
    returns, return_totals = build_return_rows(source_wb, partner_codes)
    customer_orders, customer_totals = build_customer_order_rows(source_wb)

    long_payable = (
        purchase_totals["opening"]
        + purchase_totals["counted_purchase"]
        - payment_totals["counted_payment"]
        - return_totals["counted_return"]
    )
    totals = {
        "long_opening_balance": purchase_totals["opening"],
        "long_counted_purchase": purchase_totals["counted_purchase"],
        "long_counted_payment": payment_totals["counted_payment"],
        "long_counted_return": return_totals["counted_return"],
        "long_payable": long_payable,
        "long_historical_paid": payment_totals["historical_paid"],
        "long_reference_only": purchase_totals["reference_only"] + payment_totals["reference_only"] + return_totals["reference_only"],
        "customer_order_rows": customer_totals["rows"],
        "customer_order_total": customer_totals["total"],
        "customer_order_paid": customer_totals["paid"],
        "customer_legacy_missing_price_rows": customer_totals["missing_price_rows"],
    }
    assert_expected(totals)

    wb = Workbook()
    del wb[wb.sheetnames[0]]
    write_sheet(wb, "Đối tác", PARTNER_HEADERS, partners)
    write_sheet(wb, "Nhập hàng", PURCHASE_HEADERS, purchases, {10, 12, 13})
    write_sheet(wb, "Thanh toán", PAYMENT_HEADERS, payments, {5})
    write_sheet(wb, "Trả hàng", RETURN_HEADERS, returns, {8, 9})
    write_sheet(wb, "Đơn khách", CUSTOMER_ORDER_HEADERS, customer_orders, {6, 7, 8})
    write_sheet(wb, "Đối soát", RECONCILIATION_HEADERS, build_reconciliation_rows(totals), {3})

    wb["Nhập hàng"].column_dimensions["F"].width = 38
    wb["Nhập hàng"].column_dimensions["Q"].width = 28
    wb["Đơn khách"].column_dimensions["E"].width = 36
    wb["Đơn khách"].column_dimensions["J"].width = 34
    wb["Đối soát"].column_dimensions["B"].width = 38
    wb["Đối soát"].column_dimensions["D"].width = 56

    add_dropdown(wb["Nhập hàng"], ["Có", "Không"], "O2:O500")
    add_dropdown(wb["Thanh toán"], ["Có", "Không"], "G2:G500")
    add_dropdown(wb["Trả hàng"], ["Có", "Không"], "J2:J500")
    add_dropdown(wb["Đơn khách"], ["Đã có giá", "Quên giá", "Chưa rõ"], "I2:I500")

    wb.save(OUTPUT_WORKBOOK)
    print(f"Wrote {OUTPUT_WORKBOOK.relative_to(ROOT)}")
    print(f"Đơn khách meaningful rows: {totals['customer_order_rows']}")
    print(f"Long payable: {totals['long_payable']}")
    print(f"Partner ledger rows: {len(purchases) + len(payments) + len(returns)}")
    return OUTPUT_WORKBOOK


if __name__ == "__main__":
    build_workbook()
