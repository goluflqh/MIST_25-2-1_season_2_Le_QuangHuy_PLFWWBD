import argparse
import datetime as dt
import json
import shutil
import subprocess
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_WORKBOOK = ROOT / "operations" / "minhhong-cong-no-doi-tac-copy-2026-05-26.xlsx"
DEFAULT_JSON = ROOT / "output" / "minhhong-legacy-import.json"


def cell_value(value):
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    return value


def records(ws):
    headers = [cell.value for cell in ws[1]]
    result = []
    for source_row, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(value not in (None, "") for value in row):
            continue
        result.append({headers[index]: cell_value(value) for index, value in enumerate(row) if index < len(headers)})
        result[-1]["__sourceRow"] = source_row
    return result


def convert_workbook(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    purchases = []
    for row in records(wb["Nhập hàng"]):
        purchases.append({
            "code": row.get("Mã nhập"),
            "date": row.get("Ngày nhập"),
            "debtPartner": row.get("Đối tác công nợ"),
            "sourceName": row.get("Người bán/nguồn gốc"),
            "description": row.get("Tên hàng"),
            "category": row.get("Loại"),
            "quantity": row.get("Số lượng"),
            "unit": row.get("Đơn vị"),
            "unitPrice": row.get("Đơn giá"),
            "amount": row.get("Thành tiền"),
            "receivedGoods": row.get("Đã nhận hàng"),
            "countsInDebt": row.get("Tính công nợ"),
            "sourceRow": row.get("__sourceRow"),
        })

    payments = []
    for row in records(wb["Thanh toán"]):
        payments.append({
            "code": row.get("Mã thanh toán"),
            "date": row.get("Ngày"),
            "partner": row.get("Đối tác"),
            "amount": row.get("Số tiền"),
            "paymentMethod": row.get("Phương thức"),
            "notes": row.get("Ghi chú"),
            "countsInDebt": row.get("Tính công nợ"),
            "sourceRow": row.get("__sourceRow"),
        })

    returns = []
    for row in records(wb["Trả hàng"]):
        returns.append({
            "code": row.get("Mã trả"),
            "date": row.get("Ngày"),
            "partner": row.get("Đối tác"),
            "description": row.get("Tên hàng"),
            "category": row.get("Loại"),
            "quantity": row.get("Số lượng"),
            "unitPrice": row.get("Đơn giá"),
            "amount": row.get("Thành tiền"),
            "countsInDebt": row.get("Tính công nợ"),
            "notes": row.get("Lý do/Ghi chú"),
            "sourceRow": row.get("Dòng gốc") or row.get("__sourceRow"),
        })

    customer_orders = []
    for row in records(wb["Đơn khách"]):
        customer_orders.append({
            "code": row.get("Mã đơn"),
            "date": row.get("Ngày mua"),
            "customerName": row.get("Tên khách"),
            "customerPhone": row.get("Số điện thoại"),
            "productName": row.get("Sản phẩm"),
            "totalAmount": row.get("Tổng tiền"),
            "paidAmount": row.get("Đã thu"),
            "debtAmount": row.get("Còn nợ"),
            "notes": row.get("Ghi chú"),
            "dataStatus": row.get("Trạng thái dữ liệu"),
            "sourceRow": row.get("Dòng gốc") or row.get("__sourceRow"),
        })

    return {
        "purchases": purchases,
        "payments": payments,
        "returns": returns,
        "customerOrders": customer_orders,
    }


def main():
    parser = argparse.ArgumentParser(description="Convert/import Minh Hồng legacy workbook rows into the web database.")
    parser.add_argument("workbook", nargs="?", default=str(DEFAULT_WORKBOOK))
    parser.add_argument("--json-out", default=str(DEFAULT_JSON))
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    workbook_path = Path(args.workbook).resolve()
    json_path = Path(args.json_out).resolve()
    json_path.parent.mkdir(parents=True, exist_ok=True)

    data = convert_workbook(workbook_path)
    json_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote normalized workbook JSON: {json_path}")

    npx = shutil.which("npx.cmd") or shutil.which("npx")
    if not npx:
        raise RuntimeError("Cannot find npx/npx.cmd on PATH.")

    command = [npx, "tsx", "scripts/import-minhhong-legacy-json.ts", str(json_path)]
    if args.dry_run:
        command.append("--dry-run")
    subprocess.run(command, cwd=ROOT, check=True)


if __name__ == "__main__":
    main()
